
# -*- coding: utf-8 -*-
from django.db import models
from django import forms
from django.core.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _
from django.utils.html import format_html, format_html_join
from django.utils import timezone
import re
import os
import shutil
import datetime
import zipfile
from pathlib import Path
from openpyxl import load_workbook, cell

# Create your models here.

BASE_DIR = Path(__file__).resolve().parent.parent
MEDIA_ROOT  = BASE_DIR / 'media'

EXAM_TYPE_CHOICES = [
    ('1','C语言'),
    ('2','C++'),
    ('3','Java'),
    ('4','Python'),
]

PERIOD_CHOICES = [
    ('1','120分钟'),
    ('2','90分钟'),
]

ANSWER_CHOICES = [
    ('1','1'),
    ('2','2'),
    ('3','3'),
    ('4','4'),
]


class Student(models.Model):
    class Meta:
        verbose_name = '考试-考生'
        verbose_name_plural = '考试-考生'
    class_name = models.CharField('班级',max_length=100, default='')
    student_name = models.CharField('姓名', max_length=20, default='')
    student_id = models.CharField('学号', max_length=20,  unique = True, default='')
    pass_word = models.CharField('密码', max_length=20, default='', blank=True)

    def __str__(self):
        return self.class_name+" "+self.student_name +" "+self.student_id


class Exam(models.Model):
    class Meta:
        verbose_name = '考试-考场'
        verbose_name_plural = '考试-考场'

    pub_date = models.DateTimeField('创建时间', 'date published', null=True, default=timezone.now)
    problem_type = models.CharField("试卷类型", max_length=20, choices=EXAM_TYPE_CHOICES, default='1')
    creator = models.CharField('创建人', max_length=200, default='..老师')
    info_text = models.CharField('考试信息', max_length=200, default='C语言期末考试')
    period = models.CharField("考试时长", max_length=5, choices=PERIOD_CHOICES, default='2')

    passwd_second_login = models.CharField('二次登录密码', max_length=200, default='3333')
    opened = models.BooleanField("考场开放？", default=True)

    choice_question_num = models.IntegerField(verbose_name="选择题个数", default=20)
    choice_question_score = models.IntegerField(verbose_name="选择题分值", default=2)
    complete_question_num = models.IntegerField(verbose_name="填空题个数", default=5)
    complete_question_score = models.IntegerField(verbose_name="填空题分值", default=10)
    coding_question_num = models.IntegerField(verbose_name="编程题个数", default=1)
    coding_question_score = models.IntegerField(verbose_name="编程题分值", default=10)

    def __str__(self):
        return '考场-'+str(self.id)

    def id_(self):
        return '考场-'+str(self.id)
    id_.short_description = '考试编号'

    def clean(self):
        score = self.choice_question_num*self.choice_question_score + \
            self.coding_question_num * self.coding_question_score + self.complete_question_score * self.complete_question_num
        if score != 100:
            raise ValidationError(_('总分值不等于100'))

    def all_question_stat_(self):
        return '选择题'+str(self.choice_question_score)+'分X'+str(self.choice_question_num)+ \
              ' + ' + '填空题'+str(self.complete_question_score)+'分X'+str(self.complete_question_num)+ \
            ' + '+'编程题'+str(self.coding_question_score)+'分X'+str(self.coding_question_num)
    all_question_stat_.short_description = '考题统计'

    def exam_type_(self):
        return str(EXAM_TYPE_CHOICES[int(self.problem_type)-1][1])
    exam_type_.short_description = '考试类型'

    def period_(self):
        return str(PERIOD_CHOICES[int(self.period)-1][1])
    period_.short_description = '考试时长'

    def out_link_(self):
        return format_html('<a href="/c/examroom/{}" target="_blank">点击查看</a>'.format(self.id))
    out_link_.short_description = '考场详情'




class ExamPaper(models.Model):
    class Meta:
        verbose_name = '考试-试卷'
        verbose_name_plural = '考试-试卷'

    def __str__(self):
        return '试卷-'+str(self.unique_key)

    problem_type = models.CharField("试卷类型", max_length=20, choices=EXAM_TYPE_CHOICES, default='1')
    student = models.ForeignKey(
        Student,
        on_delete=models.CASCADE, null=True,
        verbose_name='所属考生'
    ) 
    exam = models.ForeignKey(
        Exam,
        on_delete=models.CASCADE, null=True,
        verbose_name='所属考场'
    ) 
    unique_key = models.CharField('unique-key', max_length=20, primary_key=True, default='xxx')  #  student_id_local+exam_id_local

    start_time = models.DateTimeField('开考时间', null=True, blank=True, default=timezone.now)
    end_time = models.DateTimeField('交卷时间', null=True, blank=True, default=timezone.now)
    add_time = models.IntegerField("附加延时[分]", default=0)
    enabled = models.BooleanField("是否可以作答?", default=True)
    choice_question_finished = models.BooleanField("选择题是否结束作答?", default=False)

    student_id_local = models.CharField('学号', max_length=20, default='')
    exam_id_local = models.CharField('考场号', max_length=20, default='')

    choice_questions = models.TextField("选择题列表", max_length=1000,  blank=True, default='')
    choice_question_answers = models.TextField("选择题答案列表", max_length=1000, blank=True,  default='')
    choice_question_results = models.TextField("选择题评分", max_length=1000, blank=True,  default='')

    complete_questions = models.TextField("填空题列表", max_length=1000,  blank=True, default='')
    complete_question_answers = models.TextField("填空题答案列表", max_length=1000, blank=True,  default='')
    complete_question_results = models.TextField("填空题评分", max_length=1000, blank=True,  default='')

    coding_questions = models.TextField("编程题列表", max_length=1000, blank=True, default='')
    coding_question_answers = models.TextField("编程题答案列表", max_length=1000, blank=True, default='')
    coding_question_results = models.TextField("编程题评分", max_length=1000, blank=True, default='')


    def is_ended_(self):
        if self.enabled:
            return "否"
        else:
            return "是"

    def is_empty_(self):
        if self.exam.choice_question_num:
            if len(self.choice_questions)<1: return True

        if self.exam.coding_question_num:
            if len(self.coding_questions)<1: return True
        return False


    def disable_(self):
        self.enabled = False
        self.end_time = datetime.datetime.now()
        self.save()

    def add_time_enable_(self, t=3):
        self.enabled = True
        self.add_time = self.add_time + t
        self.save()

    def start_time_(self):
        return (self.start_time.strftime("%Y-%m-%d %H:%M:%S"))
    start_time_.short_description = '开考时间'

    def end_time_(self):
        return (self.end_time.strftime("%Y-%m-%d %H:%M:%S"))
    end_time_.short_description = '结束时间'

    def coding_question_answers_(self):
        return self.coding_question_answers.split(',')
    coding_question_answers_.short_description = '编程题答案'

    def complete_question_answers_(self):
        return self.complete_question_answers.split('\n')

    def choice_question_answers_(self):
        return self.choice_question_answers.split(',')
    choice_question_answers_.short_description = '选择题答案'

    def choice_questions_all_(self):
        choice_question_ids = [int(x) for x in self.choice_questions.split(',') if len(x)]
        choice_questions_all = []
        for i in choice_question_ids:
            choice_questions_all.append(ChoiceQuestion.objects.get(pk=i))
        return choice_questions_all
    choice_questions_all_.short_description = '全部选择题'

    def complete_questions_all_(self):
        question_ids = [int(x) for x in self.complete_questions.split(',') if len(x)]
        questions_all = []
        for i in choice_question_ids:
            choice_questions_all.append(CompleteQuestion.objects.get(pk=i))
        return questions_all

    def coding_questions_all_(self):
        coding_question_ids = [int(x) for x in self.coding_questions.split(',') if len(x)]
        coding_questions_all = []
        for i in coding_question_ids:
            coding_questions_all.append(CodingQuestion.objects.get(pk=i))
        return coding_questions_all
    coding_questions_all_.short_description = '全部编程题'

    ## question_id start from 1 to n
    def choice_questions_pk_(self, question_id):
        question_database_id = int(self.choice_questions.split(',')[question_id-1])
        return ChoiceQuestion.objects.get(pk=question_database_id)
    choice_questions_pk_.short_description = 'pk选择题'

    def complete_questions_pk_(self, question_id):
        question_database_id = int(self.complete_questions.split(',')[question_id-1])
        return CompleteQuestion.objects.get(pk=question_database_id)

    def coding_questions_pk_(self, question_id):
        question_database_id = int(self.coding_questions.split(',')[question_id-1])
        return CodingQuestion.objects.get(pk=question_database_id)
    coding_questions_pk_.short_description = 'pk编程题'

    def coding_output_path_(self, question_id):
        output_save_path = os.path.join(MEDIA_ROOT, 'upload_output','coding_output_{}_{}.txt'.format(self.unique_key, question_id))
        return output_save_path
    coding_output_path_.short_description = '上传结果保存目录'

    def update_choice_question_answer_result_(self, question_id, choice_id):
        old_answers = self.choice_question_answers.split(',')
        old_answers[question_id-1] = str(choice_id)
        self.choice_question_answers = ','.join(old_answers)

        choice_question = self.choice_questions_pk_(question_id)
        if choice_question.answer == str(choice_id):
            score = '1'
        else:
            score = '0'
        old_answers = self.choice_question_results.split(',')
        old_answers[question_id-1] = score
        self.choice_question_results = ','.join(old_answers)
        self.save()

    def update_complete_question_answer_result_(self, question_id, submit_answers):
        old_answers = self.complete_question_answers.split('\n')
        old_answers[question_id-1] = ','.join(submit_answers)
        self.complete_question_answers = '\n'.join(old_answers)

        question = self.complete_questions_pk_(question_id)
        old_results = self.complete_question_results.split(',')
        old_results[question_id-1] = str(question.score(submit_answers))
        self.complete_question_results = ','.join(old_results)
        self.save()

    def choice_question_result_stat(self):
        answers = [x for x in self.choice_question_answers.split(',') if x!='+']
        results = [x for x in self.choice_question_results.split(',') if x=='1']
        return self.exam.choice_question_num, len(answers), len(results)

        # print(self.choice_question_results)

    def update_coding_question_answer_result_(self, coding_question_id, output_save_path):
        old_answers = self.coding_question_answers.split(',')
        old_answers[coding_question_id-1] = str(output_save_path)
        self.coding_question_answers = ','.join(old_answers)

        ## 
        coding_question = self.coding_questions_pk_(coding_question_id)
        answer_path = coding_question.upload_answer_file.path
        with open(answer_path, encoding='utf-8') as f:
            answers = [x.strip() for x in f.readlines()]
        with open(output_save_path, encoding='utf-8') as f:
            outputs = [x.strip() for x in f.readlines()]
        # print(outputs)
        min_len = min(len(answers), len(outputs))
        correct_cnt = 0
        for i in range(min_len):
            print(answers[i], outputs[i])
            if answers[i] == outputs[i]: correct_cnt = correct_cnt + 1
        score = str(correct_cnt*1.0/len(answers))
        old_answers = self.coding_question_results.split(',')
        old_answers[coding_question_id-1] = score
        print(old_answers)
        self.coding_question_results = ','.join(old_answers)
        self.save()

    def coding_question_result_stat(self):
        answers = [x for x in self.coding_question_answers.split(',') if x!='+']
        results = [float(x) for x in self.coding_question_results.split(',')]
        sum = 0
        for i in range(len(results)):
            sum = sum + results[i]
        return self.exam.coding_question_num, len(answers), sum

    def coding_question_result_detail(self):
        return [int(float(x)*self.exam.coding_question_score) for x in self.coding_question_results.split(',')]

    def complete_question_result_detail(self):
        return [int(float(x)*self.exam.complete_question_score) for x in self.complete_question_results.split(',')]

    def choice_question_result_detail(self):
        return [int(float(x)*self.exam.choice_question_score) for x in self.choice_question_results.split(',')]

    def total_score(self):
        scores = self.choice_question_result_detail() + self.coding_question_result_detail() + self.complete_question_result_detail()
        score = 0
        for s in scores: score = score + s
        return score


def validate_zipfile(value):
    extension = value.name.split('.')[-1]
    if extension not in ['zip',]:
        raise ValidationError(
            _(value.name+'不是zip压缩文件'),
            params={'value': value.name},
        )

def validate_txtfile(value):
    extension = value.name.split('.')[-1]
    if extension not in ['txt',]:
        raise ValidationError(
            _(value.name+'不是txt文件'),
            params={'value': value.name},
        )

def validate_cfile(value):
    extension = value.name.split('.')[-1]
    if extension not in ['c',]:
        raise ValidationError(
            _(value.name+'不是c文件'),
            params={'value': value.name},
        )

def validate_csvfile(value):
    extension = value.name.split('.')[-1]
    if extension not in ['csv',]:
        raise ValidationError(
            _(value.name+'不是csv文件'),
            params={'value': value.name},
        )



class ChoiceQuestion(models.Model):
    class Meta:
        verbose_name = '题目-选择题'
        verbose_name_plural = '题目-选择题'

    def __str__(self):
        return '选择题-'+str(self.id)

    problem_type = models.CharField("试卷类型", max_length=20, choices=EXAM_TYPE_CHOICES, default='1')

    question_text = models.TextField('题干')
    choice_1 = models.CharField('选项1', max_length=200,  default='')
    choice_2 = models.CharField('选项2', max_length=200,  default='')
    choice_3 = models.CharField('选项3', max_length=200,  default='')
    choice_4 = models.CharField('选项4', max_length=200,  default='')

    answer = models.CharField("正确选项", max_length=2, choices=ANSWER_CHOICES, default='1')

    def question_html_(self):
        question_text = [x for x in self.question_text.split('\n')]
        return format_html_join(
                '', '<p style="color:{};">{}</p>',
                (('black', x) for x in question_text)
                )
    question_html_.short_description = '题目'

    def answer_list_(self):
        choice_list = [
            ['white', 'A. '+ self.choice_1], 
            ['white', 'B. '+  self.choice_2], 
            ['white', 'C. '+  self.choice_3], 
            ['white', 'D. '+  self.choice_4], 
            ]
        choice_list[int(self.answer)-1][0] = 'Lime'

        return format_html("<ul>") + \
                format_html_join(
                '\n', '<li style="background-color:{};">{}</li>',
                ((x[0], x[1]) for x in choice_list)
                ) \
                + format_html("</ul>")
    answer_list_.short_description = '题目选项'



class CompleteQuestion(models.Model):
    class Meta:
        verbose_name = '题目-填空题'
        verbose_name_plural = '题目-填空题'

    def __str__(self):
        return '填空题-'+str(self.id)

    problem_type = models.CharField("试卷类型", max_length=20, choices=EXAM_TYPE_CHOICES, default='1')

    question_text = models.TextField('题干', help_text = '包含题目说明和代码段，其中填空数固定为3，用带圈的数字标出')
    answers = models.TextField('答案', help_text ='3个答案用英文逗号隔开')


    def score(self, submit_answers):
        answers = self.answers.split(',')
        print(answers, submit_answers)
        cnt = 0
        for i in range(len(submit_answers)):
            if submit_answers[i] == answers[i]: cnt = cnt + 1
        return  cnt*1.0 / len(answers)


    def question_html_(self):
        question_text = [x for x in self.question_text.split('\n')]
        return format_html_join(
                '', '<p style="color:{};">{}</p>',
                (('black', x) for x in question_text)
                )
    question_html_.short_description = '题目'


    def answers_html_(self):
        question_text = [x for x in self.answers.split('\n')]
        return format_html_join(
                '', '<p style="color:{};">{}</p>',
                (('black', x) for x in question_text)
                )
    answers_html_.short_description = '答案'


class CodingQuestion(models.Model):
    class Meta:
        verbose_name = '题目-编程题'
        verbose_name_plural = '题目-编程题'

    def __str__(self):
        return str(self.id)+"-"+self.question_text

    problem_type = models.CharField("试卷类型", max_length=20, choices=EXAM_TYPE_CHOICES, default='1')
    question_text = models.TextField('题目简介', )
    # upload_zipfile = models.FileField(upload_to='upload_zipfile/', null=True, blank=True, 
    # validators=[validate_zipfile], verbose_name='上传题目文件zip压缩包')

    upload_description_file = models.FileField(upload_to='upload_c_file/', null=True, blank=True, 
    validators=[validate_txtfile], verbose_name='上传题目描述[.txt文件]')

    upload_c_file = models.FileField(upload_to='upload_c_file/', null=True, blank=True, 
    validators=[validate_cfile], verbose_name='上传C文件[.c文件]')

    upload_input_file = models.FileField(upload_to='upload_answer_file/', null=True, blank=True, 
    validators=[validate_txtfile], verbose_name='上传题目输入[.txt文件]')

    upload_answer_file = models.FileField(upload_to='upload_answer_file/', null=True, blank=True, 
    validators=[validate_txtfile], verbose_name='上传题目输出[.txt文件]')

    def question_html_(self):
        try:
            with open(self.upload_description_file.path,'r', encoding='utf-8') as f:
                lines = f.readlines()
                return format_html_join(
                        '', '<p style="color:{};">{}</p>',
                        (('black', x) for x in lines)
                        )
        except:
            return ''
    question_html_.short_description = '题目'

    def code_html_(self):
        try:
            with open(self.upload_c_file.path,'r', encoding='utf-8') as f:
                lines = f.readlines()
                return format_html_join(
                        '', '<p style="color:{};">{}</p>',
                        (('black', x) for x in lines)
                        )
        except:
            return ''
    code_html_.short_description = '代码'

    def zip_path_(self):
        try:
            c_path, input_path = self.upload_c_file.path, self.upload_input_file.path
            paths = os.path.split(input_path)
            return os.path.sep.join([paths[0], 'coding-'+str(self.id)+'.zip'])
        except:
            return ''
    zip_path_.short_description = '压缩文件'

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)  
        zip_path = self.zip_path_()
        if zip_path:
            if os.path.exists(zip_path): os.remove(zip_path)
            c_path, input_path = self.upload_c_file.path, self.upload_input_file.path
            print(zip_path)
            zf = zipfile.ZipFile(zip_path, 'w')
            zf.write(c_path,'{}/main.cpp'.format('coding'))
            zf.write(input_path,'{}/input.txt'.format('coding'))
            zf.close()


Student_Type_CHOICES = (
    ("师院一本","师院一本"),
    ("南岳学院","南岳学院"),
)

class StudentInfoImporter(models.Model):
    class Meta:
        verbose_name = '考试-批量导入考生'
        verbose_name_plural = '考试-批量导入考生'

    def __str__(self):
        return '导入考生-'+str(self.id)

    upload_description_file = models.FileField(upload_to='upload_student_list/', null=True, blank=True, verbose_name='上传考生信息文件')
    student_type = models.CharField('学生来源', max_length=20, choices= Student_Type_CHOICES, default = "师院一本")

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)  # Call the "real" save() method.
        print(self.upload_description_file, 'saved')
        if self.student_type == "师院一本":
            with open(self.upload_description_file.path,'r', encoding='utf-8') as f:
                lines = [x.strip().replace('\t',' ').split(' ') for x in f.readlines()[:] if len(x)>0]
                class_name = ''
                for x in lines:
                    if len(x)==1: 
                        class_name = x[0]
                        continue
                    if len(x)>5 and (x[0] !='学号'):
                        Student.objects.get_or_create(class_name=class_name, student_id=x[0], student_name=x[1])

        elif self.student_type == "南岳学院":
            wb = load_workbook(self.upload_description_file.path)
            ws = wb.active 
            print('sheetnames: ', wb.sheetnames)

            prep = ['A','B', 'J']
            row_id = 1
            empty_rows = 0
            while empty_rows<10:
                row_id = row_id + 1
                id = ws[prep[0]+str(row_id)].value
                name = ws[prep[1]+str(row_id)].value
                classname = ws[prep[2]+str(row_id)].value
                if id is None or id.strip() == '':
                    empty_rows = empty_rows +1
                    continue
                
                empty_rows = 0
                print(id, name, classname)
                if id == "学号" and name == "姓名" and classname == "班级":
                    break

            empty_rows = 0
            while empty_rows<10:
                row_id = row_id + 1
                student_id = ws[prep[0]+str(row_id)].value
                name = ws[prep[1]+str(row_id)].value
                classname = ws[prep[2]+str(row_id)].value
                if student_id is None or student_id.strip() == '' or name is None or classname is None:
                    empty_rows = empty_rows +1
                    continue
                
                empty_rows = 0
                # print(student_id, name, classname)
                Student.objects.get_or_create(class_name=classname, student_id=student_id, student_name=name)


            
def validate_xlsx(value):

    extension = value.name.split('.')[-1]
    if extension != 'xlsx':
        raise ValidationError(
            _(value.name+'不是xlsx文件'),
            params={'value': value.name},
        )



class ChoiceQuestionImporter(models.Model):
    class Meta:
        verbose_name = '题目-批量导入选择题'
        verbose_name_plural = '题目-批量导入选择题'

    def __str__(self):
        return '导入选择题'+str(self.id)

    pub_date = models.DateTimeField('导入时间', null=True, default=timezone.now)
    upload_description_file = models.FileField(upload_to='upload_import_files/', null=True, blank=True, 
    validators=[validate_xlsx], verbose_name='上传选择文件[.xlsx]')

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs) 
        print(self.upload_description_file, 'saved')

        wb = load_workbook(self.upload_description_file.path)
        ws = wb.active 
        print('sheetnames: ', wb.sheetnames)
        print('ws1: ', 
        ws['B1'].value, 
        ws['C1'].value, 
        ws['D1'].value, 
        ws['E1'].value,  
        ws['F1'].value,  
        ws['G1'].value,)
        if  ws['B1'].value != '题干' or\
        ws['C1'].value != '选项1' or\
        ws['D1'].value != '选项2' or\
        ws['E1'].value != '选项3' or\
        ws['F1'].value != '选项4' or\
        ws['G1'].value != '正确选项':
            print('导入错误：文件内容格式不对！')
            return

        prep = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
        row_id = 1
        empty_rows = 0
        while empty_rows<10:
            row_id = row_id + 1
            question_text = ws[prep[0]+str(row_id)].value
            if question_text is None or question_text.strip() == '':
                empty_rows = empty_rows +1
                continue
            
            empty_rows = 0

            choice_1 = ws[prep[1]+str(row_id)].value
            choice_2 = ws[prep[2]+str(row_id)].value
            choice_3 = ws[prep[3]+str(row_id)].value
            choice_4 = ws[prep[4]+str(row_id)].value
            answer = ws[prep[5]+str(row_id)].value

            if choice_1 and choice_2 and choice_3 and choice_4 and answer:
                print(question_text, choice_1, choice_2, choice_3, choice_4)

                ChoiceQuestion.objects.get_or_create(question_text = question_text,
                                                    choice_1=choice_1, 
                                                    choice_2=choice_2, 
                                                    choice_3=choice_3, 
                                                    choice_4=choice_4, 
                                                    answer= str(answer))

